from speedruncomapi import Game, Leaderboard
import time, datetime
import asyncio
import aiohttp
import openpyxl as pxl

start = time.perf_counter()

async def fetch_api(session, url):
    try:
        async with session.get(url, timeout=10) as response:
            response.raise_for_status()
            data = await response.json()
            return data

    except Exception as e:
        return ''

async def fetch_all_apis(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_api(session, url) for url in urls]
        return await asyncio.gather(*tasks)

def get_rest_api_data(urls):
    return asyncio.run(fetch_all_apis(urls))

if __name__ == "__main__":
    print("setting up variables...")
    def LISTMAKER_1():
        global id_list, name_list, category_id_list, beast_capture_count_list
        id_list = ['7dg6l4gkgq7dd6yq4qyoow7l', '7dg6l4gkgq7dd6yqmlnmmod1', '7dg6l4gkgq7dd6yq810jjd5l', '7dg6l4gkgq7dd6yq9qjnn8g1', '7dg6l4gk21gdd3m14qyoow7l', '7dg6l4gk21gdd3m1mlnmmod1', '7dg6l4gk21gdd3m1810jjd5l', '7dg6l4gk21gdd3m19qjnn8g1', '7dg6l4gkjqzeemmq4qyoow7l', '7dg6l4gkjqzeemmqmlnmmod1', '7dg6l4gkjqzeemmq810jjd5l', '7dg6l4gkjqzeemmq9qjnn8g1', '7dg6l4gk21dggej14qyoow7l', '7dg6l4gk21dggej1mlnmmod1', '7dg6l4gk21dggej1810jjd5l', '7dg6l4gk21dggej19qjnn8g1', '7dg6l4gk5q8889kq4qyoow7l', '7dg6l4gk5q8889kqmlnmmod1', '7dg6l4gk5q8889kq810jjd5l', '7dg6l4gk5q8889kq9qjnn8g1', '7dg6l4gk4qykjm314qyoow7l', '7dg6l4gk4qykjm31mlnmmod1', '7dg6l4gk4qykjm31810jjd5l', '7dg6l4gk4qykjm319qjnn8g1', '7dg6l4gkmlngz7614qyoow7l', '7dg6l4gkmlngz761mlnmmod1', '7dg6l4gkmlngz761810jjd5l', '7dg6l4gkmlngz7619qjnn8g1', '7dg6l4gkmln2grjl4qyoow7l', '7dg6l4gkmln2grjlmlnmmod1', '7dg6l4gkmln2grjl810jjd5l', '7dg6l4gkmln2grjl9qjnn8g1', '7dg6l4gk810r23214qyoow7l', '7dg6l4gk810r2321mlnmmod1', '7dg6l4gk810r2321810jjd5l', '7dg6l4gk810r23219qjnn8g1', '7dg6l4gklmo8vvj14qyoow7l', '7dg6l4gklmo8vvj1mlnmmod1', '7dg6l4gklmo8vvj1810jjd5l', '7dg6l4gklmo8vvj19qjnn8g1', '7dg6l4gk1w4ezzoq4qyoow7l', '7dg6l4gk1w4ezzoqmlnmmod1', '7dg6l4gk1w4ezzoq810jjd5l', '7dg6l4gk1w4ezzoq9qjnn8g1', '7dg6l4gkq8kezwyq4qyoow7l', '7dg6l4gkq8kezwyqmlnmmod1', '7dg6l4gkq8kezwyq810jjd5l', '7dg6l4gkq8kezwyq9qjnn8g1', '7dg6l4gkqyzx4g614qyoow7l', '7dg6l4gkqyzx4g61mlnmmod1', '7dg6l4gkqyzx4g61810jjd5l', '7dg6l4gkqyzx4g619qjnn8g1', '7dg6l4gkln8y3jol4qyoow7l', '7dg6l4gkln8y3jolmlnmmod1', '7dg6l4gkln8y3jol810jjd5l', '7dg6l4gkln8y3jol9qjnn8g1', '7dg6l4gkqox3k64q4qyoow7l', '7dg6l4gkqox3k64qmlnmmod1', '7dg6l4gkqox3k64q810jjd5l', '7dg6l4gkqox3k64q9qjnn8g1', '7dg6l4gk1395zjk14qyoow7l', '7dg6l4gk1395zjk1mlnmmod1', '7dg6l4gk1395zjk1810jjd5l', '7dg6l4gk1395zjk19qjnn8g1', '7dg6l4gkq75kpor14qyoow7l', '7dg6l4gkq75kpor1mlnmmod1', '7dg6l4gkq75kpor1810jjd5l', '7dg6l4gkq75kpor19qjnn8g1', '7dg6l4gkqkekv59q4qyoow7l', '7dg6l4gkqkekv59qmlnmmod1', '7dg6l4gkqkekv59q810jjd5l', '7dg6l4gkqkekv59q9qjnn8g1', '7dg6l4gkqkejr3nq4qyoow7l', '7dg6l4gkqkejr3nqmlnmmod1', '7dg6l4gkqkejr3nq810jjd5l', '7dg6l4gkqkejr3nq9qjnn8g1', 'n2yg87724lxoo6rlgq7ddxyq', 'n2yg87724lxoo6rl21gddpm1', 'n2yg87724lxoo6rljqzee3mq', 'n2yg87724lxoo6rl21dgzjj1', 'n2yg8772814ggwj1gq7ddxyq', 'n2yg8772814ggwj121gddpm1', 'n2yg8772814ggwj1jqzee3mq', 'n2yg8772814ggwj121dgzjj1', 'n2yg8772z19yypklgq7ddxyq', 'n2yg8772z19yypkl21gddpm1', 'n2yg8772z19yypkljqzee3mq', 'n2yg8772z19yypkl21dgzjj1', 'n2yg877281poo5kqgq7ddxyq', 'n2yg877281poo5kq21gddpm1', 'n2yg877281poo5kqjqzee3mq', 'n2yg877281poo5kq21dgzjj1', 'n2yg8772xqk00ny1gq7ddxyq', 'n2yg8772xqk00ny121gddpm1', 'n2yg8772xqk00ny1jqzee3mq', 'n2yg8772xqk00ny121dgzjj1', 'n2yg87728102wyjqgq7ddxyq', 'n2yg87728102wyjq21gddpm1', 'n2yg87728102wyjqjqzee3mq', 'n2yg87728102wyjq21dgzjj1', 'n2yg87729qjpxo0qgq7ddxyq', 'n2yg87729qjpxo0q21gddpm1', 'n2yg87729qjpxo0qjqzee3mq', 'n2yg87729qjpxo0q21dgzjj1', 'n2yg87729qj5pj31gq7ddxyq', 'n2yg87729qj5pj3121gddpm1', 'n2yg87729qj5pj31jqzee3mq', 'n2yg87729qj5pj3121dgzjj1', 'n2yg8772jq6j83j1gq7ddxyq', 'n2yg8772jq6j83j121gddpm1', 'n2yg8772jq6j83j1jqzee3mq', 'n2yg8772jq6j83j121dgzjj1', 'n2yg8772qoxmkkpqgq7ddxyq', 'n2yg8772qoxmkkpq21gddpm1', 'n2yg8772qoxmkkpqjqzee3mq', 'n2yg8772qoxmkkpq21dgzjj1', 'n2yg8772139wddd1gq7ddxyq', 'n2yg8772139wddd121gddpm1', 'n2yg8772139wddd1jqzee3mq', 'n2yg8772139wddd121dgzjj1', 'n2yg8772q65478vlgq7ddxyq', 'n2yg8772q65478vl21gddpm1', 'n2yg8772q65478vljqzee3mq', 'n2yg8772q65478vl21dgzjj1', 'n2yg8772lmoxn781gq7ddxyq', 'n2yg8772lmoxn78121gddpm1', 'n2yg8772lmoxn781jqzee3mq', 'n2yg8772lmoxn78121dgzjj1', 'n2yg87721w4m86mqgq7ddxyq', 'n2yg87721w4m86mq21gddpm1', 'n2yg87721w4m86mqjqzee3mq', 'n2yg87721w4m86mq21dgzjj1', 'n2yg8772139mdny1gq7ddxyq', 'n2yg8772139mdny121gddpm1', 'n2yg8772139mdny1jqzee3mq', 'n2yg8772139mdny121dgzjj1', 'n2yg8772qvvej05qgq7ddxyq', 'n2yg8772qvvej05q21gddpm1', 'n2yg8772qvvej05qjqzee3mq', 'n2yg8772qvvej05q21dgzjj1', 'n2yg87721gny56olgq7ddxyq', 'n2yg87721gny56ol21gddpm1', 'n2yg87721gny56oljqzee3mq', 'n2yg87721gny56ol21dgzjj1', 'n2yg87721pyw92e1gq7ddxyq', 'n2yg87721pyw92e121gddpm1', 'n2yg87721pyw92e1jqzee3mq', 'n2yg87721pyw92e121dgzjj1', 'n2yg8772q75jn3v1gq7ddxyq', 'n2yg8772q75jn3v121gddpm1', 'n2yg8772q75jn3v1jqzee3mq', 'n2yg8772q75jn3v121dgzjj1', 'qj70zmeq10vz5nwllmokx3j1', 'q6504o3l10vz5nwllmokx3j1', 'qj70zmeq10vz5nwl1w4pm2oq', 'q6504o3l10vz5nwl1w4pm2oq', 'qj70zmeq10vz5nwlqoxpyepq', 'q6504o3l10vz5nwlqoxpyepq', 'qj70zmeq10vz5nwl1398vkd1', 'q6504o3l10vz5nwl1398vkd1'] 
        name_list = ['Survivor Facility_0 One Player', 'Survivor Facility_0 Two Players', 'Survivor Facility_0 Three Players', 'Survivor Facility_0 Four Players', 'Survivor Airport One Player', 'Survivor Airport Two Players', 'Survivor Airport Three Players', 'Survivor Airport Four Players', 'Survivor Homestead One Player', 'Survivor Homestead Two Players', 'Survivor Homestead Three Players', 'Survivor Homestead Four Players', 'Survivor Abandoned Prison Maximus One Player', 'Survivor Abandoned Prison Maximus Two Players', 'Survivor Abandoned Prison Maximus Three Players', 'Survivor Abandoned Prison Maximus Four Players', 'Survivor Library One Player', 'Survivor Library Two Players', 'Survivor Library Three Players', 'Survivor Library Four Players', 'Survivor Abandoned Facility Remake One Player', 'Survivor Abandoned Facility Remake Two Players', 'Survivor Abandoned Facility Remake Three Players', 'Survivor Abandoned Facility Remake Four Players', 'Survivor Forgotten Facility One Player', 'Survivor Forgotten Facility Two Players', 'Survivor Forgotten Facility Three Players', 'Survivor Forgotten Facility Four Players', 'Survivor Haunted Mansion One Player', 'Survivor Haunted Mansion Two Players', 'Survivor Haunted Mansion Three Players', 'Survivor Haunted Mansion Four Players', 'Survivor The Backrooms One Player', 'Survivor The Backrooms Two Players', 'Survivor The Backrooms Three Players', 'Survivor The Backrooms Four Players', 'Survivor School One Player', 'Survivor School Two Players', 'Survivor School Three Players', 'Survivor School Four Players', 'Survivor Nuclear Power Plant One Player', 'Survivor Nuclear Power Plant Two Players', 'Survivor Nuclear Power Plant Three Players', 'Survivor Nuclear Power Plant Four Players', 'Survivor Sewer Treatment Plant One Player', 'Survivor Sewer Treatment Plant Two Players', 'Survivor Sewer Treatment Plant Three Players', 'Survivor Sewer Treatment Plant Four Players', 'Survivor Laboratory Complex One Player', 'Survivor Laboratory Complex Two Players', 'Survivor Laboratory Complex Three Players', 'Survivor Laboratory Complex Four Players', 'Survivor Abandoned Facility Optimus One Player', 'Survivor Abandoned Facility Optimus Two Players', 'Survivor Abandoned Facility Optimus Three Players', 'Survivor Abandoned Facility Optimus Four Players', 'Survivor Arcade One Player', 'Survivor Arcade Two Players', 'Survivor Arcade Three Players', 'Survivor Arcade Four Players', 'Survivor Arctic Station One Player', 'Survivor Arctic Station Two Players', 'Survivor Arctic Station Three Players', 'Survivor Arctic Station Four Players', 'Survivor Sanctuary Zoo One Player', 'Survivor Sanctuary Zoo Two Players', 'Survivor Sanctuary Zoo Three Players', 'Survivor Sanctuary Zoo Four Players', 'Survivor Spaceship One Player', 'Survivor Spaceship Two Players', 'Survivor Spaceship Three Players', 'Survivor Spaceship Four Players', 'Survivor Ghost Town One Player', 'Survivor Ghost Town Two Players', 'Survivor Ghost Town Three Players', 'Survivor Ghost Town Four Players', 'Beast Facility_0 One Capture', 'Beast Facility_0 Two Captures', 'Beast Facility_0 Three Captures', 'Beast Facility_0 Four Captures', 'Beast Airport One Capture', 'Beast Airport Two Captures', 'Beast Airport Three Captures', 'Beast Airport Four Captures', 'Beast Homestead One Capture', 'Beast Homestead Two Captures', 'Beast Homestead Three Captures', 'Beast Homestead Four Captures', 'Beast Abandoned Prison Maximus One Capture', 'Beast Abandoned Prison Maximus Two Captures', 'Beast Abandoned Prison Maximus Three Captures', 'Beast Abandoned Prison Maximus Four Captures', 'Beast Library One Capture', 'Beast Library Two Captures', 'Beast Library Three Captures', 'Beast Library Four Captures', 'Beast Abandoned Facility Remake One Capture', 'Beast Abandoned Facility Remake Two Captures', 'Beast Abandoned Facility Remake Three Captures', 'Beast Abandoned Facility Remake Four Captures', 'Beast Forgotten Facility One Capture', 'Beast Forgotten Facility Two Captures', 'Beast Forgotten Facility Three Captures', 'Beast Forgotten Facility Four Captures', 'Beast Haunted Mansion One Capture', 'Beast Haunted Mansion Two Captures', 'Beast Haunted Mansion Three Captures', 'Beast Haunted Mansion Four Captures', 'Beast The Backrooms One Capture', 'Beast The Backrooms Two Captures', 'Beast The Backrooms Three Captures', 'Beast The Backrooms Four Captures', 'Beast School One Capture', 'Beast School Two Captures', 'Beast School Three Captures', 'Beast School Four Captures', 'Beast Nuclear Power Plant One Capture', 'Beast Nuclear Power Plant Two Captures', 'Beast Nuclear Power Plant Three Captures', 'Beast Nuclear Power Plant Four Captures', 'Beast Sewer Treatment Plant One Capture', 'Beast Sewer Treatment Plant Two Captures', 'Beast Sewer Treatment Plant Three Captures', 'Beast Sewer Treatment Plant Four Captures', 'Beast Laboratory Complex One Capture', 'Beast Laboratory Complex Two Captures', 'Beast Laboratory Complex Three Captures', 'Beast Laboratory Complex Four Captures', 'Beast Abandoned Facility Optimus One Capture', 'Beast Abandoned Facility Optimus Two Captures', 'Beast Abandoned Facility Optimus Three Captures', 'Beast Abandoned Facility Optimus Four Captures', 'Beast Arcade One Capture', 'Beast Arcade Two Captures', 'Beast Arcade Three Captures', 'Beast Arcade Four Captures', 'Beast Arctic Station One Capture', 'Beast Arctic Station Two Captures', 'Beast Arctic Station Three Captures', 'Beast Arctic Station Four Captures', 'Beast Sanctuary Zoo One Capture', 'Beast Sanctuary Zoo Two Captures', 'Beast Sanctuary Zoo Three Captures', 'Beast Sanctuary Zoo Four Captures', 'Beast Spaceship One Capture', 'Beast Spaceship Two Captures', 'Beast Spaceship Three Captures', 'Beast Spaceship Four Captures', 'Beast Ghost Town One Capture', 'Beast Ghost Town Two Captures', 'Beast Ghost Town Three Captures', 'Beast Ghost Town Four Captures', 'Survivor Backrooms (Halloween) One Player', 'Beast Backrooms (Halloween) One Capture', 'Survivor Backrooms (Halloween) Two Players', 'Beast Backrooms (Halloween) Two Captures', 'Survivor Backrooms (Halloween) Three Players', 'Beast Backrooms (Halloween) Three Captures', 'Survivor Backrooms (Halloween) Four Players', 'Beast Backrooms (Halloween) Four Captures']
        category_id_list = ['7dg6l4gk', 'n2yg8772', 'wk65l5e2', 'qj70zmeq', 'q6504o3l']
        beast_capture_count_list = [['gq7ddxyq', '21gddpm1', 'jqzee3mq', '21dgzjj1'],['lmokx3j1', '1w4pm2oq', 'qoxpyepq', '1398vkd1']]
        #print(len(id_list), len(name_list))
    print("done setting up variables", f"({time.perf_counter() - start}s)")
    LISTMAKER_1()

    print("getting all runs information...")
    page_count = 1 + 2460 #Remember to change this everytime you update the board if necessary and offset by 20 ; link https://www.speedrun.com/api/v1/runs?game=46wpg3dr&offset=2460
    while True:
        placeholder_list = []
        api_endpoints = [f"https://www.speedrun.com/api/v1/runs?game=46wpg3dr&offset={i}" for i in range(20, page_count, 20)]
        all_runs = get_rest_api_data(api_endpoints)
        if '' not in all_runs: break
        else: print(f"There are error data(s) in the list, trying again... ({time.perf_counter() - start}s)")
    print("done getting all runs", f"({time.perf_counter() - start}s)")

    print("setting up correct runs' data and getting users' id...")
    placeholder_list = []
    for i in all_runs:
        for j in i["data"]:
            placeholder_list.append(j)
    placeholder_list_2 = []
    for i in range(0, len(placeholder_list)):
        run = placeholder_list[i]
        if run["status"]["status"] == "verified" and run["category"] in category_id_list:
            run.update({"place" : i})
            placeholder_list_2.append(run)
    all_runs = placeholder_list_2
    #print(all_runs, len(all_runs))
    player_id_list = []
    guest_list = []

    def GET_PLAYER_ID(list):
        global player_id_list, guest_list
        for i in range(0, len(list)) :
            if list[i]["rel"] == 'user' :
                player_id_list.append(list[i]["id"])
            else :
                guest_list.append(list[i]["name"])
    
    for i in all_runs :
        players = i["players"]
        GET_PLAYER_ID(players)
        player_id_list = list(set(player_id_list))
        guest_list = list(set(guest_list))
    print("done setting up and getting ids", f"({time.perf_counter() - start}s)")

    print("getting unique username information...")
    player_name_list = []
    player_nation_list = []
    while True:
        placeholder_list = []
        api_endpoints = [f"https://www.speedrun.com/api/v1/users/{i}" for i in player_id_list]
        data_list = get_rest_api_data(api_endpoints)
        if "" not in data_list: break            
        else: print(f"There are error data(s) in the list, trying again... ({time.perf_counter() - start}s)")

    for item in data_list:
        player_name_list.append(item["data"]["names"]["international"])
        if item["data"]["location"] == None :
            player_nation_list.append("None")
        else :
            player_nation_list.append(item["data"]["location"]["country"]["names"]["international"])
    #print(len(player_id_list), len(player_name_list))
    print("done unique username info", f"({time.perf_counter() - start}s)")

print("getting run count in all categories...")
categories_placement_count = []

def GET_CATEGORY_ID(list, category):
    category_id = ""
    if category in category_id_list:
        if category_id_list.index(category) == 0:
            category_id = category_id_list[0] + list[0] + list[1]
            category = 0
        elif category_id_list.index(category) == 1:
            category_id = category_id_list[1] + list[0] + list[1]
            category = beast_capture_count_list[0].index(list[1]) + 1
        else:
            category_id = list[0] + list[1] + list[2]
            if category_id_list.index(list[0]) == 3:
                category = 0
            else:
                category = beast_capture_count_list[1].index(list[2]) + 1
    return category_id, category

def GET_TOTAL_RUNS_IN_CATEGORY(id):
    global categories_placement_count
    for i in categories_placement_count:
        if id in i:
            i[id] += 1
            placement = i[id]
            break
    else:
        categories_placement_count.append({f"{id}": 1})
        placement = 1
    return placement

for run in all_runs :
    run["values"], run["category"] = GET_CATEGORY_ID(list(run["values"].values()), run["category"])
    run["place"] = GET_TOTAL_RUNS_IN_CATEGORY(run["values"])
#print(categories_placement_count, len(categories_placement_count))
#print(all_runs)
print("done getting run count", f"({time.perf_counter() - start}s)")

print("setting up run to desired run format...")
run_stats_datas = []
special_case = ["SkittlesCat", "Jenna_0134", "[ph]svxsul"]
K1, K2 = 150, 20
for run in all_runs:
    ftf_run = {"name" : "name","runid" : "id","gametype": 0,"time" : 900,"placement" : 0,"total_run_in_category" : 0,"the_runners" : [],"random_count" : 0,"wr_val" : 0,"wr_value_per_person" : 0,"score" : 0,"link" : "link"}
    count = 0
    ftf_run['name'] = name_list[id_list.index(run["values"])]
    ftf_run['link'] = run["weblink"]
    ftf_run['runid'] = run["id"]
    ftf_run['time'] = run["times"]["primary_t"]
    ftf_run['placement'] = run["place"]
    if run["category"] > 0:
        ftf_run["gametype"] = 1
        ftf_run['random_count'] = run["category"]
    else:
        ftf_run['random_count'] = 0
    for i in categories_placement_count:
        if run["values"] in i:
            ftf_run['total_run_in_category'] = i[run["values"]]
            break
    for i in run["players"]:
        if i["rel"] == "user":
            ftf_run['the_runners'].append(player_name_list[player_id_list.index(i["id"])])
            count += 1
        else:
            if i["name"] in special_case :
                ftf_run['the_runners'].append(i["name"])
                count += 1
            else:
               ftf_run['random_count'] += 1
    run_stats_datas.append(ftf_run)
#print(run_stats_datas[17], len(run_stats_datas))
print("done setting all runs", f"({time.perf_counter() - start}s)")

print("ordering runs...")
placeholder_list = []
for i in run_stats_datas:
    placeholder_list_2 = []
    placeholder_list_2.append(i["name"])
    placeholder_list_2.append(i["runid"])
    placeholder_list_2.append(i["time"])
    placeholder_list_2.append(i["placement"])
    placeholder_list_2.append(i["total_run_in_category"])
    placeholder_list.append(placeholder_list_2)

placeholder_list.sort(key=lambda t: t[2])
#print(placeholder_list, len(placeholder_list))
placeholder_list_2 = []
for i in placeholder_list:
    if i[0] in placeholder_list_2:
        placeholder_list_2[placeholder_list_2.index(i[0]) + 1] += 1
        i[3] = placeholder_list_2[placeholder_list_2.index(i[0]) + 1]
    else:
        i[3] = 1
        placeholder_list_2 += [i[0], 1]
for i in range(0, len(placeholder_list)-1):
    if placeholder_list[i][2] == placeholder_list[i+1][2] and placeholder_list[i][0] == placeholder_list[i+1][0]:
        placeholder_list[i+1][3] = placeholder_list[i][3]
#print(placeholder_list, len(placeholder_list))
for run in run_stats_datas:
    for i in placeholder_list:
        if i[1] == run["runid"]:
           run["time"] = i[2] 
           run["placement"] = i[3]
           break
    PL = len(run['the_runners'])
    if run["placement"] == 1:
        run['wr_val'] = 1
        run['wr_value_per_person'] = (3/PL)*(1-run["gametype"]) + 3*(run["gametype"])
    RD = run['random_count']
    if run["gametype"] == 0:
        run['score'] = run['score'] = (run['total_run_in_category']/run['placement']) + ((K1 - 10*PL)/run['time']) + (RD + 1)/5
    else:
        run['score'] = (run['total_run_in_category']/run['placement'] + ((K2*RD)/run['time']))
    #print(run, run_stats_datas.index(run))
#print(run_stats_datas[91])
print("done ordering runs", f"({time.perf_counter() - start}s)")

print("setting up player own stat...")
player_stats_datas = []

for i in range(0, len(player_name_list + special_case)):
    player_stats = {"name" : "name","country" : "country","total_wr" : 0,"survivor_wr" : 0,"beast_wr" : 0,"wr_value" : 0,"score" : 0,"most_value_run" : "run_name"}
    count = 0
    if i in player_name_list:
        player_stats['name'] = player_name_list[i]
        player_stats['country'] = player_nation_list[i]
    else:
        player_stats['name'] = (player_name_list + special_case)[i]
        player_stats['country'] = "None"
    for j in run_stats_datas:
        if (player_name_list + special_case)[i] in j["the_runners"]:
            player_stats['total_wr'] += j["wr_val"]
            player_stats['wr_value'] += j['wr_value_per_person']
            if j["placement"] == 1:
                if j['gametype'] == 0: player_stats['survivor_wr'] += 1
                else: player_stats['beast_wr'] += 1
            player_stats['score'] += j["score"]
            if count < j["score"]: player_stats['most_value_run'], count = j["name"], j["score"]
    player_stats_datas.append(player_stats)

#print(player_stats_datas)
print("done setting all players", f"({time.perf_counter() - start}s)")

print("extracting datas")
placeholder_list = []
placeholder_list_2 = []
#ftf_run = {"name" : "name","runid" : "id","gametype": 0,"time" : 900,"placement" : 0,"total_run_in_category" : 0,"the_runners" : [],"random_count" : 0,"wr_val" : 0,"wr_value_per_person" : 0,"score" : 0,"link" : "link"}
for i in run_stats_datas:
    data = {"Run name" : "","Time" : 0,"Time (in seconds)" : 0,"Place" : 0,"Runners" : "","Number of Survivors / Captures" : 0,"WR Value per Runner" : 0,"Score" : 0,"Run link" : ""}
    p = ""
    data["Run name"] = i["name"]
    data['Time'] = f"{i["time"]//60}" + ":" + f"{(i["time"] - (i["time"]//60)*60):02d}"
    data['Time (in seconds)'] = i["time"]
    data["Place"] = i["placement"]
    for j in i['the_runners']:
        p += f"{j}, "
    data["Runners"] = p
    if i["gametype"] == 0:
        data["Number of Survivors / Captures"] = (len(i["the_runners"]) + i['random_count'])
    else:
        data["Number of Survivors / Captures"] = i['random_count']
    data['WR Value per Runner'] = i["wr_value_per_person"]
    data['Score'] = i["score"]
    data['Run link'] = i["link"]
    placeholder_list.append(data)

#player_stats = {"name" : "name","country" : "country","total_wr" : 0,"survivor_wr" : 0,"beast_wr" : 0,"wr_value" : 0,"score" : 0,"most_value_run" : "run_name"}
for i in player_stats_datas:
    data = {"Player Name" : "","Nationality" : "","Total WR(s)" : 0,"Survivor WR(s)" : 0,"Beast WR(s)" : 0,"WR Value" : 0,"Total Score" : 0,"Most value run" : ""}
    data['Player Name'] = i["name"]
    data['Nationality'] = i["country"]
    data['Total WR(s)'] = i["total_wr"]
    data['Survivor WR(s)'] = i["survivor_wr"]
    data['Beast WR(s)'] = i["beast_wr"]
    data['WR Value'] = i["wr_value"]
    data['Total Score'] = i["score"]
    data['Most value run'] = i["most_value_run"]
    placeholder_list_2.append(data)

wb = pxl.Workbook()
ws = wb.active
wb.create_sheet("All Runs Data")
sheet_1 = wb["All Runs Data"]
wb.create_sheet("All Players Data")
sheet_2 = wb["All Players Data"]

sheet_1.append(list(placeholder_list[0].keys()))
for run in placeholder_list:
    sheet_1.append(list(run.values()))
sheet_2.append(list(placeholder_list_2[0].keys()))
for player in placeholder_list_2:
    sheet_2.append(list(player.values()))

cdate = datetime.datetime.now()
cdatestr = cdate.strftime("%B%d%Y %I%M%S")

wb.save(f"DATA {cdatestr}.xlsx")
print("Successfully created excel file", f"({time.perf_counter() - start}s)")
